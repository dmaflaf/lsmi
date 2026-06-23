#!/usr/bin/env python3
"""LSMI - Generador completo v8 · Modo Diurno · Temporada 2026
Archivos necesarios en la carpeta:
  - LSMI_1era/2da/3era_Division_*.xlsx  (estadisticas)
  - LSMI_Torneo_*.xlsx                  (posiciones y resultados)
  - sancionados_*.xlsx                  (sancionados)
  - logo .png                           (opcional)
  - horarios.json                       (horarios de partidos - editar cada semana)
  - LSMI_Rivales_SegundaEtapa.xlsx      (proximos rivales sin horario - respaldo)
"""
import openpyxl, json, sys, os, unicodedata, base64
from datetime import datetime

def norm(s):
    return unicodedata.normalize('NFD', str(s)).encode('ascii','ignore').decode('ascii').upper()

def simplificar_nombre(nombre):
    """Normaliza un nombre de equipo para comparaciones flexibles (quita tildes, puntos, minúsculas)."""
    if not nombre: return ''
    nfd = unicodedata.normalize('NFD', str(nombre))
    sin_tildes = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
    return sin_tildes.lower().replace('.','').replace('-','').replace("'","").strip()

# Mapeo manual para nombres que difieren demasiado entre fuentes
ALIAS_EQUIPOS = {
    'boca jrs': 'Boca Jr',
    'pradera de los sauces': 'Pradera',
    'atletico racing': 'Racing',
    'sporting athlas senior': 'Athlas Senior',
    'estandar': 'Standard',
    'scorpion': 'Scorpions',
    'scorpions': 'Scorpions',
}

def construir_mapa_equipos(posiciones):
    """Construye un diccionario {nombre_simplificado: nombre_oficial} desde las tablas de posiciones."""
    mapa = {}
    for div, tabla in posiciones.items():
        for eq in tabla:
            nombre = eq['equipo']
            mapa[simplificar_nombre(nombre)] = nombre
    # Agregar aliases manuales
    for alias, oficial in ALIAS_EQUIPOS.items():
        mapa[alias] = oficial
    return mapa

def resolver_nombre(nombre, mapa):
    """Resuelve un nombre de equipo a su forma oficial. Devuelve (nombre_resuelto, fue_corregido)."""
    key = simplificar_nombre(nombre)
    if key in mapa:
        oficial = mapa[key]
        return oficial, (oficial != nombre)
    return nombre, False

def safe_print(msg):
    """Print seguro para consolas Windows que no soportan emojis."""
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode('ascii', 'replace').decode('ascii'))

HOJAS_POS = {'1era División':'1era División','2da División':'2da División','3era División C1':'3era División 1','3era División C2':'3era División 2'}
DIV_RES      = {v:k for k,v in HOJAS_POS.items()}
DIV_RIESGO   = {'1era División':4,'2da División':2,'3era División C1':4,'3era División C2':4}
DIV_CLASSIFY = {'1era División':8,'2da División':8,'3era División C1':8,'3era División C2':8}
SERIE_TO_DIV = {'A':'1era División','B':'2da División','C1':'3era División C1','C2':'3era División C2'}

def procesar_stats(ruta):
    wb=openpyxl.load_workbook(ruta,read_only=True,data_only=True)
    ws_g=wb['GOLEADORES']; goles_pj={}; uf=0
    for row in ws_g.iter_rows(min_row=3,values_only=True):
        if not row[0] or not isinstance(row[0],(int,float)): continue
        e=str(row[1]).strip() if row[1] else ''; j=str(row[2]).strip() if row[2] else ''; g=row[3]
        if e and j and g and isinstance(g,(int,float)):
            key=(e,j); goles_pj[key]=goles_pj.get(key,0)+int(g); uf=max(uf,int(row[0]))
    ws_a=wb['AMONESTACIONES']; tarj_pj={}
    for row in ws_a.iter_rows(min_row=3,values_only=True):
        if not row[0] or not isinstance(row[0],(int,float)): continue
        e=str(row[1]).strip() if row[1] else ''; j=str(row[2]).strip() if row[2] else ''; t=str(row[3]).strip() if row[3] else ''
        if e and j and t:
            key=(e,j)
            if key not in tarj_pj: tarj_pj[key]={'amarillas':0,'dobles':0,'rojas':0}
            if 'Doble' in t or 'doble' in t: tarj_pj[key]['dobles']+=1
            elif 'Roja' in t or 'roja' in t or '🔴' in t: tarj_pj[key]['rojas']+=1
            else: tarj_pj[key]['amarillas']+=1
    equipos=sorted(set(k[0] for k in goles_pj)|set(k[0] for k in tarj_pj))
    teams_data={}
    for eq in equipos:
        ge={j:g for (e,j),g in goles_pj.items() if e==eq}
        te={j:t for (e,j),t in tarj_pj.items() if e==eq}
        tg=sum(ge.values()); tam=sum(t['amarillas'] for t in te.values())
        tdo=sum(t['dobles'] for t in te.values()); tro=sum(t['rojas'] for t in te.values())
        pg=round(tg/uf,2) if uf else 0; pt=round((tam+tdo+tro)/uf,2) if uf else 0
        ts=sorted(ge.items(),key=lambda x:-x[1])
        tcr=[{'nombre':j,'amarillas':t['amarillas'],'dobles':t['dobles'],'rojas':t['rojas'],'_p':t['amarillas']+t['dobles']*2+t['rojas']*3} for j,t in te.items() if t['amarillas']+t['dobles']*2+t['rojas']*3>0]
        tcr.sort(key=lambda x:-x['_p'])
        teams_data[eq]={'total_goles':tg,'promedio_goles':pg,'total_amarillas':tam,'total_doble':tdo,'total_rojas':tro,'promedio_tarjetas':pt,
            'top_scorers':[{'nombre':j,'goles':g} for j,g in ts],
            'top_cards':[{k:v for k,v in c.items() if k!='_p'} for c in tcr]}
    all_s=[{'nombre':j,'equipo':e,'goles':g} for (e,j),g in goles_pj.items()]; all_s.sort(key=lambda x:-x['goles'])
    all_c=[{'nombre':j,'equipo':e,'amarillas':t['amarillas'],'dobles':t['dobles'],'rojas':t['rojas'],'_p':t['amarillas']+t['dobles']*2+t['rojas']*3} for (e,j),t in tarj_pj.items() if t['amarillas']+t['dobles']*2+t['rojas']*3>0]
    all_c.sort(key=lambda x:-x['_p'])
    fp=[{'equipo':eq,'total_tarjetas':sum(t['amarillas']+t['dobles']+t['rojas'] for t in {j:t for (e,j),t in tarj_pj.items() if e==eq}.values())} for eq in equipos]
    fp.sort(key=lambda x:x['total_tarjetas'])
    return {'meta':{'ultima_fecha':uf,'generado':datetime.now().strftime('%d/%m/%Y %H:%M')},
        'globalData':{'top_10_goleadores':all_s[:10],'top_10_amonestados':[{k:v for k,v in c.items() if k!='_p'} for c in all_c[:10]],'fair_play':fp},
        'teamsData':teams_data,
        '_goles_raw':{f"{e}||{j}":g for (e,j),g in goles_pj.items()},
        '_tarjetas_raw':{f"{e}||{j}":t for (e,j),t in tarj_pj.items()}}

def leer_posiciones(wb_pos):
    pos={}
    for div,hoja in HOJAS_POS.items():
        if hoja not in wb_pos.sheetnames: continue
        ws=wb_pos[hoja]; tabla=[]; ok=False
        for row in ws.iter_rows(values_only=True):
            if row[0]=='POS' and row[1]=='EQUIPO': ok=True; continue
            if not ok: continue
            if row[0] and isinstance(row[0],(int,float)) and row[1] and isinstance(row[1],str) and len(str(row[1]).strip())>1:
                if 'PJ=' in str(row[1]) or 'Puntuación' in str(row[1]): break
                tabla.append({'pos':int(row[0]),'equipo':str(row[1]).strip(),'pj':int(row[2] or 0),'pg':int(row[3] or 0),'pe':int(row[4] or 0),'pp':int(row[5] or 0),'gf':int(row[6] or 0),'gc':int(row[7] or 0),'dg':int(row[8] or 0),'pts':int(row[9] or 0)})
        pos[div]=tabla
    return pos

def leer_resultados(wb_pos):
    ws_r=wb_pos['RESULTADOS']; res={}
    for row in ws_r.iter_rows(min_row=2,values_only=True):
        if not row[0] or not isinstance(row[0],(int,float)): continue
        if not row[2] or not row[5]: continue
        fecha=int(row[0]); div_raw=str(row[1]).strip() if row[1] else ''
        local=str(row[2]).strip(); gf=int(row[3] or 0); gv=int(row[4] or 0); vis=str(row[5]).strip()
        if not div_raw or not local or not vis: continue
        div_nombre=DIV_RES.get(div_raw,div_raw)
        if div_nombre not in res: res[div_nombre]={}
        if fecha not in res[div_nombre]: res[div_nombre][fecha]=[]
        res[div_nombre][fecha].append({'local':local,'gl':gf,'gv':gv,'visitante':vis})
    return {div:[{'fecha':f,'partidos':ps} for f,ps in sorted(fechas.items(),reverse=True)] for div,fechas in res.items()}

def leer_sancionados(ruta):
    DIV_MAP={norm('1era DIVISIÓN'):'1era División',norm('2da DIVISIÓN'):'2da División',norm('3era DIVISIÓN 1'):'3era División C1',norm('3era DIVISIÓN 2'):'3era División C2'}
    wb=openpyxl.load_workbook(ruta,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]; sanc={}; div_actual=None
    for row in ws.iter_rows(values_only=True):
        c1,c2,c3,c4,c5=row[1],row[2],row[3],row[4],row[5]
        if c1 is None and c2 and isinstance(c2,str):
            c2n=norm(c2)
            for k,v in DIV_MAP.items():
                if k in c2n: div_actual=v; sanc.setdefault(div_actual,[]); break
            continue
        if div_actual and c2 and isinstance(c2,str) and c3 and isinstance(c3,str) and c4 and isinstance(c4,str):
            if norm(c2) in ('NOMBRES Y APELLIDOS','SANCION'): continue
            numero=str(c1).strip() if c1 is not None else '—'
            s_upper=str(c4).upper()
            if any(x in s_upper for x in ['INDEFINIDO','AÑO','AÑOS','MESES','PROFESIONAL']): tipo='grave'
            elif any(x in s_upper for x in ['2 FECHA','3 FECHA']): tipo='media'
            else: tipo='leve'
            fase2_str = c5.strftime('%d/%m/%Y') if hasattr(c5,'strftime') else str(c5).strip() if c5 else ''
            sanc[div_actual].append({'numero':numero,'nombre':str(c2).strip(),'club':str(c3).strip(),'sancion':str(c4).strip(),'fase2':fase2_str,'tipo':tipo})
    return sanc

def leer_horarios_json(carpeta):
    """Lee el JSON de horarios exportado del planificador (formato liga_sanmiguel_backup)."""
    DIAS_ES={'Mon':'Lun','Tue':'Mar','Wed':'Mié','Thu':'Jue','Fri':'Vie','Sat':'Sáb','Sun':'Dom'}
    MESES_ES={'Jan':'Ene','Feb':'Feb','Mar':'Mar','Apr':'Abr','May':'May','Jun':'Jun','Jul':'Jul','Aug':'Ago','Sep':'Sep','Oct':'Oct','Nov':'Nov','Dec':'Dic'}
    candidatos=[f for f in os.listdir(carpeta) if f.lower().endswith('.json') and ('backup' in f.lower() or f.lower()=='horarios.json')]
    if not candidatos: return []
    ruta=os.path.join(carpeta, candidatos[0])
    try:
        with open(ruta,encoding='utf-8') as f: raw=json.load(f)
    except Exception: return []
    partidos=[]
    for item in raw.get('data',[]):
        serie=item.get('serie','')
        div=SERIE_TO_DIV.get(serie)
        if not div: continue
        fecha_iso=item.get('fecha','')
        try:
            dt=datetime.strptime(fecha_iso,'%Y-%m-%d')
            dia_en=dt.strftime('%a'); mes_en=dt.strftime('%b')
            dia_fmt=f"{DIAS_ES.get(dia_en,dia_en)} {dt.day} {MESES_ES.get(mes_en,mes_en)}"
        except Exception:
            dia_fmt=fecha_iso
        partidos.append({
            'division':div,'fecha_iso':fecha_iso,'dia':dia_fmt,
            'hora':item.get('horarios','Sin definir'),
            'cancha':item.get('escenario','Sin definir'),
            'local':item.get('local','').strip(),
            'visitante':item.get('visitante','').strip(),
            'fase':item.get('fase',''),
            'veedor':item.get('veedor','').strip()
        })
    partidos.sort(key=lambda p:(p['fecha_iso'],p['hora']))
    return partidos

def leer_rivales_proximos(carpeta):
    """Lee Excel de próximos rivales (sin horario aún) como respaldo."""
    ruta=next((os.path.join(carpeta,f) for f in os.listdir(carpeta) if 'rivales' in f.lower() and f.endswith('.xlsx')),None)
    if not ruta: return {}
    wb=openpyxl.load_workbook(ruta,read_only=True,data_only=True)
    rivales={}
    for div_nombre,hoja in HOJAS_POS.items():
        if hoja not in wb.sheetnames: continue
        ws=wb[hoja]; rows=list(ws.iter_rows(values_only=True))
        if not rows: continue
        header=rows[0]
        fecha_cols=[(i,h) for i,h in enumerate(header) if h and str(h).startswith('F') and h!='F1 (Jugada)']
        eq_riv={}
        for row in rows[1:]:
            if not row[0]: continue
            equipo=str(row[0]).strip(); rivs=[]
            for idx,fname in fecha_cols:
                if row[idx]: rivs.append({'fecha_label':fname,'rival':str(row[idx]).strip()})
            eq_riv[equipo]=rivs
        rivales[div_nombre]=eq_riv
    return rivales

def buscar_excels(carpeta):
    archivos=[f for f in os.listdir(carpeta) if f.endswith('.xlsx')]
    stats={}
    claves={'1era División':['1era','primera'],'2da División':['2da','segunda'],'3era División C1':['division_1','3era_division_1','_1_v','g1','c1'],'3era División C2':['division_2','3era_division_2','_2_v','g2','c2']}
    for arch in archivos:
        low=arch.lower()
        if 'rivales' in low: continue
        for div,pals in claves.items():
            if any(p in low for p in pals) and div not in stats: stats[div]=os.path.join(carpeta,arch); break
    pos_ruta=next((os.path.join(carpeta,f) for f in archivos if any(k in f.lower() for k in ['torneo','posicion','fase']) and 'rivales' not in f.lower()),None)
    sanc_ruta=next((os.path.join(carpeta,f) for f in archivos if 'sancionado' in f.lower()),None)
    logo_ruta=next((os.path.join(carpeta,f) for f in os.listdir(carpeta) if f.lower().endswith('.png')),None)
    return stats,pos_ruta,sanc_ruta,logo_ruta

def procesar_todo(carpeta):
    stats_rutas,pos_ruta,sanc_ruta,logo_ruta=buscar_excels(carpeta)
    safe_print("[+] Archivos:"); [safe_print(f"   {d}: {os.path.basename(r)}") for d,r in stats_rutas.items()]
    if pos_ruta: safe_print(f"   Posiciones: {os.path.basename(pos_ruta)}")
    if sanc_ruta: safe_print(f"   Sancionados: {os.path.basename(sanc_ruta)}")
    if logo_ruta: safe_print(f"   Logo: {os.path.basename(logo_ruta)}")

    res_stats={}
    for nombre,ruta in stats_rutas.items():
        safe_print(f"\n>> {nombre}..."); res_stats[nombre]=procesar_stats(ruta)
        total_amon=sum(len(d['top_cards']) for d in res_stats[nombre]['teamsData'].values())
        safe_print(f"   {len(res_stats[nombre]['teamsData'])} equipos | F{res_stats[nombre]['meta']['ultima_fecha']} | {total_amon} amonestados")

    posiciones={}; res_partidos={}
    if pos_ruta:
        wb_pos=openpyxl.load_workbook(pos_ruta,read_only=True,data_only=True)
        posiciones=leer_posiciones(wb_pos); res_partidos=leer_resultados(wb_pos)

    sancionados={}
    if sanc_ruta: sancionados=leer_sancionados(sanc_ruta)

    horarios=leer_horarios_json(carpeta)
    safe_print(f"\n[+] Horarios cargados: {len(horarios)} partidos")
    rivales_prox=leer_rivales_proximos(carpeta)
    safe_print(f"[+] Rivales proximos (respaldo): {sum(len(v) for v in rivales_prox.values())} equipos")

    # ── NORMALIZACIÓN DE NOMBRES ──
    mapa_eq = construir_mapa_equipos(posiciones)
    correcciones = 0

    # Normalizar nombres en horarios
    for p in horarios:
        for campo in ['local','visitante']:
            nuevo, corregido = resolver_nombre(p[campo], mapa_eq)
            if corregido:
                safe_print(f"   [CORREGIDO horarios] '{p[campo]}' -> '{nuevo}'")
                p[campo] = nuevo
                correcciones += 1

    # Normalizar nombres en sancionados
    for div, lista in sancionados.items():
        for s in lista:
            nuevo, corregido = resolver_nombre(s['club'], mapa_eq)
            if corregido:
                safe_print(f"   [CORREGIDO sancionados] '{s['club']}' -> '{nuevo}'")
                s['club'] = nuevo
                correcciones += 1

    # Normalizar nombres en rivales próximos
    nuevos_rivales = {}
    for div, eq_riv in rivales_prox.items():
        nuevos_rivales[div] = {}
        for equipo, rivs in eq_riv.items():
            eq_nuevo, _ = resolver_nombre(equipo, mapa_eq)
            nuevos_rivs = []
            for r in rivs:
                riv_nuevo, _ = resolver_nombre(r['rival'], mapa_eq)
                nuevos_rivs.append({'fecha_label':r['fecha_label'],'rival':riv_nuevo})
            nuevos_rivales[div][eq_nuevo] = nuevos_rivs
    rivales_prox = nuevos_rivales

    # Enriquecer horarios con resultado si el partido ya fue jugado (busca por local+visitante en la misma división)
    for p in horarios:
        div = p['division']
        if div in res_partidos:
            for fd in res_partidos[div]:
                for partido in fd['partidos']:
                    if partido['local'] == p['local'] and partido['visitante'] == p['visitante']:
                        p['gl'] = partido['gl']
                        p['gv'] = partido['gv']
                        break

    if correcciones:
        safe_print(f"\n[OK] {correcciones} nombres de equipos corregidos automaticamente.")
    else:
        safe_print("\n[OK] Todos los nombres de equipos coinciden con la tabla de posiciones.")

    logo_b64=''
    if logo_ruta:
        with open(logo_ruta,'rb') as f: logo_b64='data:image/png;base64,'+base64.b64encode(f.read()).decode()

    ranking_eq_amon=[]
    for dn,data in res_stats.items():
        for eq,stats in data['teamsData'].items():
            tam=stats['total_amarillas']; tdo=stats['total_doble']; tro=stats['total_rojas']
            pts=tam+tdo*2+tro*3
            ranking_eq_amon.append({'equipo':eq,'division':dn,'amarillas':tam,'dobles':tdo,'rojas':tro,'puntos':pts})
    ranking_eq_amon.sort(key=lambda x:-x['puntos'])

    ultimos_por_equipo={}
    for div,fechas in res_partidos.items():
        for fd in sorted(fechas,key=lambda x:x['fecha'],reverse=True):
            for p in fd['partidos']:
                for eq in [p['local'],p['visitante']]:
                    if eq not in ultimos_por_equipo: ultimos_por_equipo[eq]=[]
                    if len(ultimos_por_equipo[eq])<5:
                        es_local=p['local']==eq
                        gf=p['gl'] if es_local else p['gv']; gc=p['gv'] if es_local else p['gl']
                        resultado='G' if gf>gc else ('E' if gf==gc else 'P')
                        ultimos_por_equipo[eq].append({'fecha':fd['fecha'],'rival':p['visitante'] if es_local else p['local'],'gf':gf,'gc':gc,'resultado':resultado,'local':es_local})

    total_goles_t=0; total_part_t=0
    for div,fechas in res_partidos.items():
        for fd in fechas:
            for p in fd['partidos']: total_goles_t+=p['gl']+p['gv']; total_part_t+=1
    prom=round(total_goles_t/total_part_t,2) if total_part_t else 0

    all_g,all_c,fp_t,stats_div=[],[],[],[]
    for dn,data in res_stats.items():
        for key,goles in data['_goles_raw'].items():
            e,j=key.split('||'); all_g.append({'nombre':j,'equipo':e,'division':dn,'goles':goles})
        for key,t in data['_tarjetas_raw'].items():
            e,j=key.split('||'); p=t['amarillas']+t['dobles']*2+t['rojas']*3
            if p>0: all_c.append({'nombre':j,'equipo':e,'division':dn,'amarillas':t['amarillas'],'dobles':t['dobles'],'rojas':t['rojas'],'_p':p})
        for eq_fp in data['globalData']['fair_play']:
            fp_t.append({'equipo':eq_fp['equipo'],'division':dn,'total_tarjetas':eq_fp['total_tarjetas']})
        tg=sum(d['total_goles'] for d in data['teamsData'].values())
        tt=sum(d['total_amarillas']+d['total_doble']+d['total_rojas'] for d in data['teamsData'].values())
        ne=len(data['teamsData'])
        stats_div.append({'division':dn,'total_goles':tg,'total_tarjetas':tt,'num_equipos':ne,'promedio_goles_equipo':round(tg/ne,1) if ne else 0})
    all_g.sort(key=lambda x:-x['goles']); all_c.sort(key=lambda x:-x['_p']); fp_t.sort(key=lambda x:x['total_tarjetas']); stats_div.sort(key=lambda x:-x['total_goles'])
    top_c=[{k:v for k,v in c.items() if k!='_p'} for c in all_c[:15]]
    global_torneo={'top_15_goleadores':all_g[:15],'top_15_amonestados':top_c,'fair_play':fp_t[:20],'ranking_divisiones':stats_div,
        'prom_goles_partido':prom,'total_partidos':total_part_t,'ranking_equipos_amonestados':ranking_eq_amon[:15],
        'total_equipos':sum(d['num_equipos'] for d in stats_div),'total_tarjetas':sum(d['total_tarjetas'] for d in stats_div),'total_goles':sum(d['total_goles'] for d in stats_div)}
    for d in res_stats.values(): d.pop('_goles_raw',None); d.pop('_tarjetas_raw',None)
    return {'generado':datetime.now().strftime('%d/%m/%Y %H:%M'),'divisiones':res_stats,'posiciones':posiciones,'resultados':res_partidos,
        'sancionados':sancionados,'globalTorneo':global_torneo,'logo':logo_b64,
        'horarios':horarios,'rivalesProximos':rivales_prox,'ultimos_por_equipo':ultimos_por_equipo}

def generar_html(datos):
    return HTML.replace('__JSON__', json.dumps(datos, ensure_ascii=False))

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>LSMI · Estadísticas Oficiales 2026</title>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800;900&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#f4f5f7;--surf:#ffffff;--surf2:#f0f2f5;--surf3:#e8ebf0;
  --border:#dde2ea;--border2:#c5ccd8;
  --red:#c8102e;--red-light:#fff0f2;--red-mid:#fbc8cf;
  --gold:#b8860b;--gold-light:#fef9e7;
  --green:#1a7a3c;--green-light:#edfaf3;
  --orange:#c45c00;--orange-light:#fff4ec;
  --blue:#1d4ed8;
  --gray:#1a1f2e;--gray2:#3d4460;--gray3:#6b7394;--gray4:#9ba3bf;
  --text:#1a1f2e;--text2:#3d4460;--text3:#6b7394;--text4:#9ba3bf;
  --shadow-sm:0 1px 4px rgba(0,0,0,.08),0 2px 8px rgba(0,0,0,.04);
  --shadow:0 4px 20px rgba(0,0,0,.10),0 1px 4px rgba(0,0,0,.06);
  --shadow-lg:0 8px 40px rgba(0,0,0,.13),0 2px 8px rgba(0,0,0,.07);
  --shadow-red:0 4px 20px rgba(200,16,46,.15),0 1px 4px rgba(200,16,46,.08);
  --1era:#c8102e;--2da:#1d4ed8;--c1:#1a7a3c;--c2:#b8860b;
}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;overflow-x:hidden;padding-bottom:80px;-webkit-font-smoothing:antialiased}

.site-header{position:fixed;top:0;width:100%;z-index:100;height:64px;background:rgba(255,255,255,.97);backdrop-filter:blur(20px);border-bottom:2px solid var(--red);box-shadow:0 2px 24px rgba(200,16,46,.12),0 1px 4px rgba(0,0,0,.06);display:flex;align-items:center;padding:0 20px;gap:14px}
.logo-img{width:44px;height:44px;object-fit:contain;border-radius:10px;flex-shrink:0;box-shadow:0 2px 10px rgba(200,16,46,.25);border:2px solid rgba(200,16,46,.15)}
.brand-title{font-family:'Barlow Condensed',sans-serif;font-size:19px;font-weight:900;letter-spacing:1px;text-transform:uppercase;color:var(--gray);line-height:1}
.brand-sub{font-size:9px;font-weight:700;color:var(--text3);letter-spacing:3px;text-transform:uppercase;margin-top:2px}

.site-nav{position:fixed;top:64px;width:100%;z-index:90;height:50px;background:linear-gradient(135deg,#1a1f2e 0%,#252b3d 100%);border-bottom:1px solid rgba(255,255,255,.08);display:flex;align-items:center;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.18)}
.nav-inner{display:flex;overflow-x:auto;white-space:nowrap;padding:0 16px;gap:2px;scrollbar-width:none;height:100%;align-items:center}
.nav-inner::-webkit-scrollbar{display:none}
.nav-btn{display:flex;align-items:center;gap:6px;padding:6px 16px;font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:rgba(255,255,255,.5);border:none;background:transparent;cursor:pointer;border-bottom:3px solid transparent;height:50px;transition:all .2s;white-space:nowrap}
.nav-btn .material-symbols-outlined{font-size:17px}
.nav-btn:hover{color:rgba(255,255,255,.85)}
.nav-btn.active{color:#fff;border-bottom-color:var(--red);background:rgba(200,16,46,.12)}

/* ── SUB-NAV (POS/RES/EST/SAN) escritorio Y móvil ── */
.sub-nav{position:sticky;top:114px;z-index:80;background:var(--surf);border-bottom:2px solid var(--border);box-shadow:var(--shadow-sm);margin:0 -16px 20px;padding:0 16px}
.sub-nav-inner{max-width:1200px;margin:0 auto;display:flex;gap:4px;overflow-x:auto;scrollbar-width:none}
.sub-nav-inner::-webkit-scrollbar{display:none}
.subn-btn{display:flex;align-items:center;gap:7px;padding:13px 18px;font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:var(--text3);border:none;background:transparent;cursor:pointer;border-bottom:3px solid transparent;white-space:nowrap;transition:all .18s}
.subn-btn .material-symbols-outlined{font-size:18px}
.subn-btn:hover{color:var(--text2);background:var(--surf2)}
.subn-btn.active{color:var(--red);border-bottom-color:var(--red);background:linear-gradient(to bottom,var(--red-light),transparent)}

.bottom-nav{position:fixed;bottom:0;width:100%;z-index:100;background:rgba(255,255,255,.97);backdrop-filter:blur(16px);border-top:1px solid var(--border);display:none;justify-content:space-around;padding:6px 0 10px;box-shadow:0 -4px 24px rgba(0,0,0,.10)}
.bnav-item{display:flex;flex-direction:column;align-items:center;gap:2px;cursor:pointer;padding:4px 12px;border-radius:8px;transition:all .2s;color:var(--text3);border:none;background:transparent}
.bnav-item.active{color:var(--red)}
.bnav-item .material-symbols-outlined{font-size:22px}
.bnav-item span:last-child{font-family:'Barlow Condensed',sans-serif;font-size:10px;font-weight:800;letter-spacing:1px;text-transform:uppercase}

.main{padding:24px 16px;max-width:1200px;margin:114px auto 0}

.view{display:none}
.view.active{display:block}
.sub-panel{display:none}
.sub-panel.active{display:block}

.s-eyebrow{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:3px;text-transform:uppercase;color:var(--red);margin-bottom:6px}
.s-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(26px,5vw,40px);font-weight:900;text-transform:uppercase;color:var(--gray);line-height:1;margin-bottom:24px}
.s-title span{color:var(--red)}

.bento-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:28px}
@media(min-width:640px){.bento-grid{grid-template-columns:repeat(3,1fr)}}
@media(min-width:1024px){.bento-grid{grid-template-columns:repeat(6,1fr)}}
.kpi-card{background:var(--surf);border:1px solid var(--border);border-radius:16px;padding:20px 16px;position:relative;overflow:hidden;transition:all .25s cubic-bezier(.4,0,.2,1);box-shadow:var(--shadow-sm)}
.kpi-card:hover{box-shadow:var(--shadow);transform:translateY(-2px)}
.kpi-card.accent{border-top:3px solid var(--red);background:linear-gradient(135deg,#fff 60%,#fff5f6 100%);box-shadow:var(--shadow-red)}
.kpi-card .live-dot{position:absolute;top:12px;right:12px;width:7px;height:7px;border-radius:50%;background:var(--red);box-shadow:0 0 8px var(--red);animation:pulse 2s infinite}
@keyframes pulse{0%{transform:scale(.95);opacity:.7}70%{transform:scale(1.1);opacity:1}100%{transform:scale(.95);opacity:.7}}
.kpi-label{font-size:9px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--text3);margin-bottom:8px;display:block}
.kpi-val{font-family:'Barlow Condensed',sans-serif;font-size:36px;font-weight:900;line-height:1;color:var(--gray)}
.kpi-val.red{color:var(--red)}
.kpi-icon{position:absolute;bottom:-8px;right:-4px;opacity:.04;font-size:72px;color:var(--gray)}

.div-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;margin-bottom:32px}
@media(min-width:768px){.div-grid{grid-template-columns:repeat(4,1fr)}}
.div-card{background:var(--surf);border:1px solid var(--border);border-radius:16px;overflow:hidden;transition:all .25s cubic-bezier(.4,0,.2,1);box-shadow:var(--shadow-sm)}
.div-card:hover{box-shadow:var(--shadow-lg);transform:translateY(-3px)}
.div-card-stripe{height:4px;width:100%}
.div-card-body{padding:16px}
.div-card-name{font-family:'Barlow Condensed',sans-serif;font-size:14px;font-weight:800;letter-spacing:.5px;text-transform:uppercase}
.div-card-eq{font-size:9px;font-weight:700;letter-spacing:1.5px;color:var(--text3);text-transform:uppercase;margin-top:2px}
.div-card-rank{font-family:'Barlow Condensed',sans-serif;font-size:36px;font-weight:900;color:var(--border2);line-height:1}
.div-card-stats{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-top:12px;padding-top:12px;border-top:1px solid var(--border)}
.div-stat{text-align:center}
.div-stat-val{font-family:'Barlow Condensed',sans-serif;font-size:18px;font-weight:800;color:var(--gray)}
.div-stat-lbl{font-size:9px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--text4);margin-top:2px}

.grid-2{display:grid;grid-template-columns:1fr;gap:20px;margin-bottom:24px}
@media(min-width:900px){.grid-2{grid-template-columns:1fr 1fr}}
.grid-3{display:grid;grid-template-columns:1fr;gap:16px;margin-bottom:24px}
@media(min-width:900px){.grid-3{grid-template-columns:1fr 1fr 1fr}}

.panel{background:var(--surf);border:1px solid var(--border);border-radius:16px;overflow:hidden;margin-bottom:20px;box-shadow:var(--shadow-sm);transition:box-shadow .2s}
.panel:hover{box-shadow:var(--shadow)}
.panel-head{display:flex;align-items:center;gap:10px;padding:14px 18px;border-bottom:1px solid var(--border);background:linear-gradient(to right,var(--surf2),var(--surf))}
.panel-head .material-symbols-outlined{font-size:20px;color:var(--red)}
.panel-head-title{font-family:'Barlow Condensed',sans-serif;font-size:14px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:var(--gray)}
.panel-body{padding:10px}

.scorer-row{display:flex;align-items:center;gap:12px;padding:10px;border-radius:10px;transition:background .15s;margin-bottom:2px}
.scorer-row:hover{background:var(--surf2)}
.scorer-row.top1{background:linear-gradient(90deg,var(--red-light) 0%,transparent 80%);border-left:3px solid var(--red);box-shadow:inset 0 0 0 1px rgba(200,16,46,.06)}
.scorer-row.top3{background:var(--surf2)}
.rank-num{font-family:'Barlow Condensed',sans-serif;font-size:15px;font-weight:900;color:var(--text4);min-width:26px;height:26px;display:inline-flex;align-items:center;justify-content:center;border-radius:6px;background:var(--surf2);text-align:center}
.rank-num.gold{color:#fff;background:linear-gradient(135deg,#d4a017,#b8860b)}
.rank-num.silver{color:#fff;background:linear-gradient(135deg,#8a9099,#6b7280)}
.rank-num.bronze{color:#fff;background:linear-gradient(135deg,#b07030,#92400e)}
.scorer-info{flex:1;min-width:0}
.scorer-name{font-size:12px;font-weight:700;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.scorer-meta{font-size:10px;font-weight:600;margin-top:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.scorer-bar-wrap{height:2px;background:var(--surf3);border-radius:2px;overflow:hidden;margin-top:5px}
.scorer-bar{height:100%;border-radius:2px}
.scorer-val{font-family:'Barlow Condensed',sans-serif;font-size:21px;font-weight:900;color:var(--gray);min-width:32px;text-align:right}

.card-t{display:inline-flex;flex-direction:column;align-items:center;justify-content:center;width:34px;height:40px;border-radius:5px;font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:900;margin-left:4px}
.card-t.roja{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
.card-t.doble{background:#fff7ed;color:#9a3412;border:1px solid #fed7aa}
.card-t.amarilla{background:#fefce8;color:#854d0e;border:1px solid #fde68a}

.fp-row{display:flex;align-items:center;gap:10px;padding:7px 10px;border-radius:8px;transition:background .15s;margin-bottom:2px}
.fp-row:hover{background:var(--surf2)}
.fp-bar-wrap{flex:1;height:4px;background:var(--surf3);border-radius:2px;overflow:hidden}
.fp-bar{height:100%;border-radius:2px}
.fp-val{font-family:'Barlow Condensed',sans-serif;font-size:14px;font-weight:800;color:var(--gray);min-width:28px;text-align:right}

.pos-wrap{overflow-x:auto;border-radius:10px;border:1px solid var(--border)}
.pos-table{width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed}
.pos-table thead{background:linear-gradient(135deg,#1a1f2e 0%,#252b3d 100%)}
.pos-table th{font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:10px 5px;text-align:center;white-space:nowrap}
.pos-table th.eq-h{text-align:left;width:120px;padding-left:10px}
.pos-table th.w30{width:28px}
.pos-table td{padding:9px 5px;text-align:center;border-bottom:1px solid var(--border);color:var(--text2);font-weight:500}
.pos-table td.eq-td{text-align:left;font-weight:700;color:var(--text);padding-left:10px;width:120px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:120px}
.pos-table tbody tr:hover td{background:var(--surf2)}
.pos-table tbody tr:last-child td{border-bottom:none}
.pts-td{font-family:'Barlow Condensed',sans-serif;font-size:16px;font-weight:900}
.dg-pos{color:var(--green);font-weight:700}
.dg-neg{color:var(--red);font-weight:700}
.pos-badge{display:inline-flex;align-items:center;justify-content:center;width:20px;height:20px;border-radius:4px;font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:900}
.pb-gold{background:var(--gold-light);color:var(--gold)}
.pb-clf{background:var(--green-light);color:var(--green)}
.pb-bot{background:var(--orange-light);color:var(--orange)}
.pb-norm{background:var(--surf2);color:var(--text3)}
.zone-clf td{background:rgba(26,122,60,.04)}
.zone-clf td.eq-td{border-left:3px solid var(--green)}
.zone-bot td{background:rgba(196,92,0,.04)}
.zone-bot td.eq-td{border-left:3px solid var(--orange)}
.zone-top td.eq-td{border-left:3px solid var(--gold)}
.tbl-legend{display:flex;gap:14px;flex-wrap:wrap;padding:10px 14px;border-top:1px solid var(--border);font-size:10px;font-weight:600;color:var(--text4);background:var(--surf2)}
.ldot{width:8px;height:8px;border-radius:2px;display:inline-block;margin-right:4px;vertical-align:middle}

.fecha-block{margin-bottom:18px}
.fecha-lbl{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:var(--text3);padding:5px 12px;background:var(--surf2);border:1px solid var(--border);border-radius:5px;margin-bottom:10px;display:inline-block}
.match-card{display:flex;align-items:center;padding:10px 14px;background:var(--surf);border:1px solid var(--border);border-radius:12px;margin-bottom:6px;gap:10px;transition:all .2s}
.match-card:hover{box-shadow:var(--shadow);transform:translateY(-1px)}
.match-team{font-size:12px;font-weight:600;color:var(--text2);flex:1;line-height:1.3}
.match-team.right{text-align:right}
.match-team.winner{color:var(--text);font-weight:700}
.match-score{font-family:'Barlow Condensed',sans-serif;font-size:19px;font-weight:900;color:var(--gray);background:linear-gradient(135deg,var(--surf2),var(--surf3));border:1px solid var(--border2);padding:5px 14px;border-radius:8px;min-width:58px;text-align:center;letter-spacing:2px;box-shadow:inset 0 1px 2px rgba(0,0,0,.05)}
.match-score.draw{color:var(--text3)}

.fixture-card{display:flex;align-items:center;padding:12px 14px;background:var(--surf);border:1px solid var(--border);border-left:4px solid var(--accent,var(--red));border-radius:12px;margin-bottom:8px;gap:10px;box-shadow:var(--shadow-sm);transition:all .2s cubic-bezier(.4,0,.2,1)}
.fixture-card:hover{box-shadow:var(--shadow);transform:translateX(2px)}
.fixture-time{display:flex;flex-direction:column;align-items:center;justify-content:center;min-width:50px;flex-shrink:0}
.fixture-hora{font-family:'Barlow Condensed',sans-serif;font-size:15px;font-weight:900;color:var(--red);line-height:1.1;white-space:nowrap}
.fixture-hora.tbd{color:var(--text4);font-size:10px}
.fixture-dia{font-size:8px;font-weight:700;color:var(--text3);letter-spacing:.5px;text-transform:uppercase;margin-top:3px;text-align:center;line-height:1.2;white-space:nowrap}
.fixture-teams{flex:1 1 auto;display:flex;align-items:center;justify-content:center;gap:8px;min-width:0;overflow:hidden}
.fixture-team{font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:800;color:var(--gray);text-align:center;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.fixture-vs{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;color:var(--text4);padding:0 2px;flex-shrink:0}
.fixture-cancha{font-size:9px;font-weight:700;color:var(--text3);letter-spacing:.5px;text-transform:uppercase;white-space:nowrap;flex-shrink:0;display:flex;align-items:center;gap:2px}
.fixture-cancha.tbd{color:var(--text4);font-style:italic}
.fixture-veedor{font-size:9px;font-weight:600;color:var(--text4);white-space:nowrap;flex-shrink:0;display:flex;align-items:center;gap:2px;margin-left:6px;padding-left:6px;border-left:1px solid var(--border)}
@media(max-width:560px){
  .fixture-card{flex-wrap:wrap;padding:10px 12px}
  .fixture-teams{order:1;width:100%;justify-content:space-between;margin-top:6px}
  .fixture-time{order:0}
  .fixture-cancha{order:2;width:100%;margin-top:6px;justify-content:center}
}

.result-pill{display:inline-flex;align-items:center;justify-content:center;width:26px;height:26px;border-radius:50%;font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:900}
.result-pill.G{background:var(--green-light);color:var(--green)}
.result-pill.E{background:var(--surf2);color:var(--text3)}
.result-pill.P{background:var(--red-light);color:var(--red)}

.sanc-table{width:100%;border-collapse:collapse;font-size:11px}
.sanc-table th{font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:10px 12px;text-align:left;background:var(--gray);white-space:nowrap}
.sanc-table td{padding:9px 12px;border-bottom:1px solid var(--border);color:var(--text2);font-size:11px}
.sanc-table tbody tr:hover td{background:var(--surf2)}
.sanc-table tbody tr:last-child td{border-bottom:none}
.sanc-name{font-weight:700;color:var(--text)}
.s-tag{display:inline-flex;align-items:center;font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;padding:2px 7px;border-radius:3px;letter-spacing:.5px;white-space:nowrap}
.s-tag.grave{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
.s-tag.media{background:#fff7ed;color:#9a3412;border:1px solid #fed7aa}
.s-tag.leve{background:#fefce8;color:#854d0e;border:1px solid #fde68a}
.s-fase2{font-family:'Barlow Condensed',sans-serif;font-size:10px;font-weight:800;color:var(--red);background:var(--red-light);border:1px solid var(--red-mid);padding:2px 7px;border-radius:3px}
.s-num{font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:900;color:var(--text3);background:var(--surf2);border:1px solid var(--border);width:28px;height:28px;border-radius:4px;display:inline-flex;align-items:center;justify-content:center}

.div-hero{background:linear-gradient(135deg,var(--surf) 70%,var(--surf2) 100%);border:1px solid var(--border);border-radius:16px;border-top:4px solid var(--accent,var(--red));padding:22px 24px;margin-bottom:20px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:14px;box-shadow:var(--shadow)}
.div-hero-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(22px,4vw,32px);font-weight:900;letter-spacing:1px;text-transform:uppercase}
.div-hero-meta{font-size:11px;color:var(--text3);font-weight:500;margin-top:3px}
.div-kpis{display:flex;gap:24px;flex-wrap:wrap}
.div-kpi-val{font-family:'Barlow Condensed',sans-serif;font-size:26px;font-weight:900;color:var(--gray)}
.div-kpi-lbl{font-size:9px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--text4)}

.team-sel-wrap{background:var(--surf);border:1px solid var(--border);border-radius:10px;padding:12px 16px;margin-bottom:18px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;box-shadow:var(--shadow-sm)}
.team-sel-lbl{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:var(--text3);white-space:nowrap}
.team-sel{flex:1;min-width:160px;background:var(--surf);color:var(--text);border:1.5px solid var(--border2);border-radius:6px;padding:8px 12px;font-family:'Barlow Condensed',sans-serif;font-size:13px;font-weight:700;cursor:pointer;outline:none}
.team-sel:focus{border-color:var(--red)}

.team-kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(100px,1fr));gap:1px;background:var(--border);border-radius:10px;overflow:hidden;margin-bottom:18px;box-shadow:var(--shadow-sm)}
.team-kpi{background:var(--surf2);padding:14px;text-align:center}
.team-kpi-val{font-family:'Barlow Condensed',sans-serif;font-size:32px;font-weight:900;color:var(--gray);line-height:1}
.team-kpi-lbl{font-size:9px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--text4);margin-top:4px}

.forma-wrap{display:flex;align-items:center;gap:6px;margin-top:8px;flex-wrap:wrap}
.forma-match{display:flex;flex-direction:column;align-items:center;gap:2px}
.forma-pill{display:inline-flex;align-items:center;justify-content:center;width:28px;height:28px;border-radius:50%;font-family:'Barlow Condensed',sans-serif;font-size:12px;font-weight:900}
.forma-pill.G{background:var(--green-light);color:var(--green);border:1px solid var(--green)}
.forma-pill.E{background:var(--surf2);color:var(--text3);border:1px solid var(--border)}
.forma-pill.P{background:var(--red-light);color:var(--red);border:1px solid var(--red)}
.forma-score{font-size:9px;color:var(--text4);font-weight:600}

.empty{color:var(--text4);font-size:12px;padding:22px;text-align:center;font-style:italic}
.section-gap{margin-bottom:28px}

.search-input{width:100%;background:var(--surf);color:var(--text);border:1.5px solid var(--border2);border-radius:8px;padding:10px 14px 10px 38px;font-family:'Inter',sans-serif;font-size:12px;font-weight:500;outline:none;transition:border-color .2s,box-shadow .2s;margin-bottom:12px;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' fill='%239ba3bf' viewBox='0 0 24 24'%3E%3Cpath d='M15.5 14h-.79l-.28-.27A6.471 6.471 0 0 0 16 9.5 6.5 6.5 0 1 0 9.5 16c1.61 0 3.09-.59 4.23-1.57l.27.28v.79l5 4.99L20.49 19l-4.99-5zm-6 0C7.01 14 5 11.99 5 9.5S7.01 5 9.5 5 14 7.01 14 9.5 11.99 14 9.5 14z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:12px center}
.search-input:focus{border-color:var(--red);box-shadow:0 0 0 3px rgba(200,16,46,.1)}
.search-input::placeholder{color:var(--text4);font-weight:400}

.site-footer{background:linear-gradient(135deg,#1a1f2e 0%,#252b3d 100%);border-top:2px solid var(--red);padding:36px 20px 100px;text-align:center;box-shadow:0 -4px 24px rgba(0,0,0,.15)}
.footer-brand{font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:3px;text-transform:uppercase;color:rgba(255,255,255,.4);margin-bottom:10px}
.footer-links{display:flex;gap:16px;flex-wrap:wrap;justify-content:center;font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.4);margin-bottom:14px}
.footer-links .nix{color:var(--red)}
.footer-quote{font-size:12px;color:rgba(255,255,255,.3);font-style:italic;max-width:480px;margin:0 auto;line-height:1.6}

@media(max-width:600px){
  .site-header{padding:0 12px}
  .brand-title{font-size:14px}
  .main{padding:16px 10px}
  .div-hero{padding:14px 16px}
  .bento-grid{gap:8px}
  .kpi-val{font-size:28px}
  .fixture-card{padding:10px 12px}
  .sub-nav{display:none}
  body.division-activa .bottom-nav{display:flex}
}
</style>
</head>
<body>

<header class="site-header">
  <img id="logoImg" class="logo-img" alt="LSMI">
  <div>
    <h1 class="brand-title">Liga San Miguel de Ibarra</h1>
    <div class="brand-sub">LSMI · Estadísticas Oficiales</div>
  </div>
</header>

<nav class="site-nav">
  <div class="nav-inner">
    <button class="nav-btn active" onclick="switchView(this,'global')"><span class="material-symbols-outlined">public</span>Global</button>
    <button class="nav-btn" onclick="switchView(this,'1era División')"><span class="material-symbols-outlined">military_tech</span>1era División</button>
    <button class="nav-btn" onclick="switchView(this,'2da División')"><span class="material-symbols-outlined">workspace_premium</span>2da División</button>
    <button class="nav-btn" onclick="switchView(this,'3era División C1')"><span class="material-symbols-outlined">star</span>3era C1</button>
    <button class="nav-btn" onclick="switchView(this,'3era División C2')"><span class="material-symbols-outlined">star_half</span>3era C2</button>
  </div>
</nav>

<main class="main">
  <div id="view-global" class="view active"></div>
  <div id="view-1era División" class="view"></div>
  <div id="view-2da División" class="view"></div>
  <div id="view-3era División C1" class="view"></div>
  <div id="view-3era División C2" class="view"></div>
</main>

<footer class="site-footer">
  <div class="footer-brand">Liga San Miguel de Ibarra · Estadísticas Oficiales</div>
  <div class="footer-links">
    <span class="nix">Powered by NIX 26</span>
  </div>
  <p class="footer-quote">"Fomentando el deporte y la disciplina en la comunidad de Ibarra. Datos proporcionados por la Comisión Técnica LSMI."</p>
</footer>

<nav class="bottom-nav" id="bottomNav">
  <button class="bnav-item active" data-sub="pos" onclick="switchSub(this,'pos')"><span class="material-symbols-outlined" style="font-variation-settings:'FILL' 1">format_list_numbered</span><span>Posiciones</span></button>
  <button class="bnav-item" data-sub="res" onclick="switchSub(this,'res')"><span class="material-symbols-outlined">calendar_today</span><span>Resultados</span></button>
  <button class="bnav-item" data-sub="stats" onclick="switchSub(this,'stats')"><span class="material-symbols-outlined">analytics</span><span>Estadísticas</span></button>
  <button class="bnav-item" data-sub="sanc" onclick="switchSub(this,'sanc')"><span class="material-symbols-outlined">gavel</span><span>Sancionados</span></button>
</nav>

<script>
const DATA = __JSON__;
const DC={'1era División':'#c8102e','2da División':'#1d4ed8','3era División C1':'#1a7a3c','3era División C2':'#b8860b'};
const DR={'1era División':4,'2da División':2,'3era División C1':4,'3era División C2':4};
const DQ={'1era División':8,'2da División':8,'3era División C1':8,'3era División C2':8};
// Estado de sub-pestaña ACTIVA, uno por división — evita que se pisen entre sí
const subState={'1era División':'pos','2da División':'pos','3era División C1':'pos','3era División C2':'pos'};
let curDiv='global';

if(DATA.logo) document.getElementById('logoImg').src=DATA.logo;

function switchView(btn,id){
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  document.querySelectorAll('.view').forEach(v=>v.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('view-'+id).classList.add('active');
  curDiv=id;
  document.title = id==='global' ? 'LSMI · Estadísticas Oficiales 2026' : id+' · LSMI';
  document.body.classList.toggle('division-activa', id!=='global');
  if(id!=='global'){
    const activeSub=subState[id]||'pos';
    document.querySelectorAll('#bottomNav .bnav-item').forEach(b=>b.classList.toggle('active',b.dataset.sub===activeSub));
    showSubPanel(id,activeSub);
  }
  window.scrollTo({top:0,behavior:'smooth'});
}

function switchSub(btn,sub){
  if(curDiv==='global') return;
  subState[curDiv]=sub;
  document.querySelectorAll('#bottomNav .bnav-item').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  showSubPanel(curDiv,sub);
}

function selectSub(divId,sub){
  subState[divId]=sub;
  document.querySelectorAll('#bottomNav .bnav-item').forEach(b=>b.classList.toggle('active',b.dataset.sub===sub));
  showSubPanel(divId,sub);
}

function showSubPanel(divId,sub){
  const p=document.getElementById('view-'+divId);
  if(!p) return;
  p.querySelectorAll('.sub-panel').forEach(sp=>sp.classList.remove('active'));
  p.querySelectorAll('.subn-btn').forEach(sb=>sb.classList.toggle('active', sb.dataset.sub===sub));
  const sp=document.getElementById('sub-'+divId+'-'+sub);
  if(sp) sp.classList.add('active');
}

function rn(i){
  const c=i===0?'gold':i===1?'silver':i===2?'bronze':'';
  return`<span class="rank-num ${c}">${i+1}</span>`;
}
function pb(pos,n,clf,rsk){
  if(pos===1) return`<span class="pos-badge pb-gold">${pos}</span>`;
  if(pos<=clf) return`<span class="pos-badge pb-clf">${pos}</span>`;
  if(pos>n-rsk) return`<span class="pos-badge pb-bot">${pos}</span>`;
  return`<span class="pos-badge pb-norm">${pos}</span>`;
}
function cardB(c){
  let b='';
  if(c.rojas>0)    b+=`<span class="card-t roja">${c.rojas}<small style="font-size:8px">R</small></span>`;
  if(c.dobles>0)   b+=`<span class="card-t doble">${c.dobles}<small style="font-size:8px">D</small></span>`;
  if(c.amarillas>0)b+=`<span class="card-t amarilla">${c.amarillas}<small style="font-size:8px">A</small></span>`;
  return b;
}

// ── GLOBAL ──
function buildGlobal(){
  const g=DATA.globalTorneo;
  const vg=document.getElementById('view-global');
  const totalEq=g.ranking_divisiones.reduce((a,d)=>a+d.num_equipos,0);
  const totalT=g.ranking_divisiones.reduce((a,d)=>a+d.total_tarjetas,0);

  const horarios=DATA.horarios||[];
  let pxHTML='<div class="empty">Sin horarios cargados aún</div>';
  if(horarios.length){
    const byDiv={};
    horarios.forEach(p=>{ if(!byDiv[p.division]) byDiv[p.division]=[]; byDiv[p.division].push(p); });
    pxHTML='';
    Object.entries(byDiv).forEach(([div,parts])=>{
      const color=DC[div]||'#c8102e';
      pxHTML+=`<div style="margin-bottom:18px">`;
      pxHTML+=`<div style="font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:${color};margin-bottom:8px;padding-left:2px">${div}</div>`;
      const now=new Date();
      parts.forEach(p=>{
        const horaTbd=p.hora==='Sin definir';
        const canchaTbd=p.cancha==='Sin definir';
        const horaFmt=horaTbd?'00:00':p.hora.replace('h',':');
        const matchDt=new Date(p.fecha_iso+'T'+horaFmt+':00');
        const isPast=now>matchDt;
        const hasResult=p.gl!==undefined&&p.gl!==null;
        if(isPast&&hasResult){
          const gl=p.gl,gv=p.gv;
          const lw=gl>gv,vw=gv>gl;
          pxHTML+=`<div class="fixture-card resultado-card" style="--accent:${color}">`;
          pxHTML+=`<div class="fixture-time"><div style="font-family:'Barlow Condensed',sans-serif;font-size:10px;font-weight:800;letter-spacing:1px;color:var(--green);background:var(--green-light);padding:2px 6px;border-radius:4px">FIN</div><div class="fixture-dia">${p.dia}</div></div>`;
          pxHTML+=`<div class="fixture-teams">
            <div class="fixture-team" style="${lw?'font-weight:900;color:'+color:''}">${p.local}</div>
            <div class="fixture-vs" style="font-family:'Barlow Condensed',sans-serif;font-size:19px;font-weight:900;color:var(--gray);background:linear-gradient(135deg,var(--surf2),var(--surf3));border:1px solid var(--border2);padding:4px 14px;border-radius:8px;min-width:58px;text-align:center;letter-spacing:2px;box-shadow:inset 0 1px 2px rgba(0,0,0,.05)">${gl} – ${gv}</div>
            <div class="fixture-team" style="${vw?'font-weight:900;color:'+color:''}">${p.visitante}</div>
          </div>`;
          pxHTML+=`<div class="fixture-cancha"><span class="material-symbols-outlined" style="font-size:13px">location_on</span>${p.cancha}</div>`;
          if(p.veedor) pxHTML+=`<div class="fixture-veedor"><span class="material-symbols-outlined" style="font-size:13px">badge</span>${p.veedor}</div>`;
          pxHTML+=`</div>`;
        } else {
          pxHTML+=`<div class="fixture-card" style="--accent:${color}">`;
          pxHTML+=`<div class="fixture-time"><div class="fixture-hora${horaTbd?' tbd':''}">${p.hora}</div><div class="fixture-dia">${p.dia}</div></div>`;
          pxHTML+=`<div class="fixture-teams"><div class="fixture-team">${p.local}</div><div class="fixture-vs">VS</div><div class="fixture-team">${p.visitante}</div></div>`;
          pxHTML+=`<div class="fixture-cancha${canchaTbd?' tbd':''}"><span class="material-symbols-outlined" style="font-size:13px">location_on</span>${p.cancha}</div>`;
          if(p.veedor) pxHTML+=`<div class="fixture-veedor"><span class="material-symbols-outlined" style="font-size:13px">badge</span>${p.veedor}</div>`;
          pxHTML+=`</div>`;
        }
      });
      pxHTML+=`</div>`;
    });
  }

  vg.innerHTML=`
  <div style="margin-bottom:22px">
    <div class="s-eyebrow">Temporada 2026</div>
    <h2 class="s-title">Estadística <span>Global</span> del Torneo</h2>
  </div>
  <div class="bento-grid">
    <div class="kpi-card accent"><div class="live-dot"></div><span class="kpi-label">Goles Totales</span><div class="kpi-val red">${g.total_goles}</div><span class="material-symbols-outlined kpi-icon">sports_soccer</span></div>
    <div class="kpi-card"><span class="kpi-label">Equipos</span><div class="kpi-val">${totalEq}</div></div>
    <div class="kpi-card"><span class="kpi-label">Divisiones</span><div class="kpi-val">${g.ranking_divisiones.length}</div></div>
    <div class="kpi-card"><span class="kpi-label">Tarjetas Totales</span><div class="kpi-val">${totalT}</div></div>
    <div class="kpi-card accent"><span class="kpi-label">Goles Líder</span><div class="kpi-val red">${g.top_15_goleadores[0]?.goles||0}</div></div>
    <div class="kpi-card"><span class="kpi-label">Prom. Goles/Partido</span><div class="kpi-val">${g.prom_goles_partido}</div></div>
  </div>
  <div class="div-grid" id="gDivCards"></div>

  <div class="panel section-gap">
    <div class="panel-head"><span class="material-symbols-outlined">event</span><span class="panel-head-title">Partidos de la Semana — Horarios & Resultados</span></div>
    <div class="panel-body">${pxHTML}</div>
  </div>

  <div class="grid-2">
    <div>
      <div class="panel">
        <div class="panel-head"><span class="material-symbols-outlined">leaderboard</span><span class="panel-head-title">Top 15 Goleadores</span></div>
        <div class="panel-body" id="gGol"></div>
      </div>
    </div>
    <div>
      <div class="panel">
        <div class="panel-head"><span class="material-symbols-outlined">warning</span><span class="panel-head-title">Ranking Amonestaciones</span></div>
        <div class="panel-body" id="gTarj"></div>
      </div>
      <div class="panel">
        <div class="panel-head"><span class="material-symbols-outlined" style="color:var(--green)">volunteer_activism</span><span class="panel-head-title">Fair Play General</span></div>
        <div class="panel-body" id="gFP"></div>
      </div>
    </div>
  </div>

  <div class="panel">
    <div class="panel-head"><span class="material-symbols-outlined">gavel</span><span class="panel-head-title">Ranking Amonestación General — Equipos</span></div>
    <div class="panel-body" style="padding:0">
      <div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12px">
        <thead style="background:var(--gray)"><tr>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:center;border-bottom:1px solid var(--border);width:40px">Pos</th>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:left;border-bottom:1px solid var(--border)">Equipo / División</th>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:center;border-bottom:1px solid var(--border);width:50px">R</th>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:center;border-bottom:1px solid var(--border);width:50px">D</th>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:center;border-bottom:1px solid var(--border);width:50px">A</th>
          <th style="font-family:'Barlow Condensed',sans-serif;font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.65);padding:11px 10px;text-align:right;border-bottom:1px solid var(--border);width:60px">Total</th>
        </tr></thead>
        <tbody id="eqAmonBody"></tbody>
      </table></div>
    </div>
  </div>`;

  const dc=document.getElementById('gDivCards');
  g.ranking_divisiones.forEach((d,i)=>{
    const color=DC[d.division]||'#c8102e';
    dc.innerHTML+=`<div class="div-card">
      <div class="div-card-stripe" style="background:${color}"></div>
      <div class="div-card-body">
        <div style="display:flex;justify-content:space-between;align-items:flex-start">
          <div><div class="div-card-name" style="color:${color}">${d.division}</div><div class="div-card-eq">${d.num_equipos} Equipos</div></div>
          <div class="div-card-rank">#${i+1}</div>
        </div>
        <div class="div-card-stats">
          <div class="div-stat"><div class="div-stat-val" style="color:${color}">${d.total_goles}</div><div class="div-stat-lbl">Goles</div></div>
          <div class="div-stat"><div class="div-stat-val">${d.promedio_goles_equipo}</div><div class="div-stat-lbl">Prom/Eq</div></div>
          <div class="div-stat"><div class="div-stat-val">${d.total_tarjetas}</div><div class="div-stat-lbl">Tarjetas</div></div>
        </div>
      </div>
    </div>`;
  });

  const maxG=g.top_15_goleadores[0]?.goles||1;
  const gg=document.getElementById('gGol');
  g.top_15_goleadores.forEach((p,i)=>{
    const color=DC[p.division]||'#c8102e', pct=Math.round(p.goles/maxG*100);
    gg.innerHTML+=`<div class="scorer-row ${i===0?'top1':i<3?'top3':''}">${rn(i)}<div class="scorer-info">
      <div class="scorer-name">${p.nombre}</div><div class="scorer-meta" style="color:${color}">${p.equipo} · ${p.division}</div>
      <div class="scorer-bar-wrap"><div class="scorer-bar" style="width:${pct}%;background:${color}"></div></div>
    </div><div class="scorer-val">${p.goles}</div></div>`;
  });

  const gt=document.getElementById('gTarj');
  g.top_15_amonestados.forEach((c,i)=>{
    const color=DC[c.division]||'#c8102e';
    gt.innerHTML+=`<div class="scorer-row">${rn(i)}<div class="scorer-info">
      <div class="scorer-name">${c.nombre}</div><div class="scorer-meta" style="color:${color}">${c.equipo} · ${c.division}</div>
    </div><div style="display:flex;gap:2px;flex-wrap:wrap">${cardB(c)}</div></div>`;
  });

  const maxFP=Math.max(...g.fair_play.map(f=>f.total_tarjetas))||1;
  const gfp=document.getElementById('gFP');
  g.fair_play.slice(0,14).forEach((eq,i)=>{
    const color=DC[eq.division]||'#c8102e', pct=Math.round(eq.total_tarjetas/maxFP*100);
    const bc=i<3?'#1a7a3c':i<8?'#b8860b':'#c8102e';
    gfp.innerHTML+=`<div class="fp-row">${rn(i)}<div class="scorer-info">
      <div class="scorer-name" style="font-size:11px">${eq.equipo}</div><div class="scorer-meta" style="color:${color}">${eq.division}</div>
      <div class="fp-bar-wrap"><div class="fp-bar" style="width:${pct}%;background:${bc}"></div></div>
    </div><div class="fp-val">${eq.total_tarjetas}</div></div>`;
  });

  const tbody=document.getElementById('eqAmonBody');
  g.ranking_equipos_amonestados.forEach((eq,i)=>{
    const color=DC[eq.division]||'#c8102e';
    tbody.innerHTML+=`<tr onmouseover="this.querySelectorAll('td').forEach(t=>t.style.background='var(--surf2)')" onmouseout="this.querySelectorAll('td').forEach(t=>t.style.background='')">
      <td style="text-align:center;padding:11px 10px;border-bottom:1px solid var(--border)"><span style="font-family:'Barlow Condensed',sans-serif;font-size:${i===0?20:16}px;font-weight:900;color:${i===0?'var(--gold)':'var(--text4)'}">${i+1}</span></td>
      <td style="padding:11px 10px;border-bottom:1px solid var(--border)">
        <div style="font-family:'Barlow Condensed',sans-serif;font-size:15px;font-weight:800;color:var(--text)">${eq.equipo}</div>
        <div style="font-size:9px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:${color};margin-top:2px">${eq.division}</div>
      </td>
      <td style="text-align:center;padding:11px 10px;border-bottom:1px solid var(--border)"><span class="card-t roja" style="width:auto;height:auto;padding:2px 8px">${eq.rojas}R</span></td>
      <td style="text-align:center;padding:11px 10px;border-bottom:1px solid var(--border)"><span class="card-t doble" style="width:auto;height:auto;padding:2px 8px">${eq.dobles}D</span></td>
      <td style="text-align:center;padding:11px 10px;border-bottom:1px solid var(--border)"><span class="card-t amarilla" style="width:auto;height:auto;padding:2px 8px">${eq.amarillas}A</span></td>
      <td style="text-align:right;padding:11px 10px;border-bottom:1px solid var(--border);font-family:'Barlow Condensed',sans-serif;font-size:20px;font-weight:900;color:${i===0?'var(--red)':'var(--gray)'}">${eq.puntos}</td>
    </tr>`;
  });
}

function buildTablaPos(divNombre){
  const tabla=DATA.posiciones[divNombre]||[];
  const color=DC[divNombre]||'#c8102e';
  const clf=DQ[divNombre]||8, rsk=DR[divNombre]||3;
  if(!tabla.length) return`<div class="empty">Sin tabla disponible</div>`;
  const n=tabla.length;
  let rows='';
  tabla.forEach(eq=>{
    const pos=eq.pos;
    const isTop=pos===1,isClf=pos>1&&pos<=clf,isBot=pos>n-rsk;
    const rowCls=isTop?'zone-top':isClf?'zone-clf':isBot?'zone-bot':'';
    const dgCls=eq.dg>0?'dg-pos':eq.dg<0?'dg-neg':'';
    rows+=`<tr class="${rowCls}">
      <td class="w30">${pb(pos,n,clf,rsk)}</td>
      <td class="eq-td">${eq.equipo}</td>
      <td class="w30 pts-td" style="color:${color}">${eq.pts}</td>
      <td class="w30 ${dgCls}">${eq.dg>0?'+':''}${eq.dg}</td>
      <td class="w30">${eq.pj}</td>
      <td class="w30">${eq.pg}</td><td class="w30">${eq.pe}</td><td class="w30">${eq.pp}</td>
      <td class="w30">${eq.gf}</td><td class="w30">${eq.gc}</td>
    </tr>`;
  });
  return`<div class="pos-wrap"><table class="pos-table">
    <thead><tr>
      <th class="w30">#</th><th class="eq-h">Equipo</th>
      <th class="w30" style="color:${color}">PTS</th><th class="w30">DG</th><th class="w30">PJ</th>
      <th class="w30">G</th><th class="w30">E</th><th class="w30">P</th><th class="w30">GF</th><th class="w30">GC</th>
    </tr></thead><tbody>${rows}</tbody></table>
    <div class="tbl-legend">
      <span><span class="ldot" style="background:var(--green)"></span>Clasificación Fase 3 (top ${clf})</span>
      <span><span class="ldot" style="background:var(--orange)"></span>Zona de riesgo (últimos ${rsk})</span>
    </div></div>`;
}

function buildResultados(divNombre){
  const fechas=DATA.resultados[divNombre]||[];
  let html=fechas.length?'':'<div class="empty">Sin resultados registrados</div>';
  fechas.forEach(f=>{
    html+=`<div class="fecha-block"><div class="fecha-lbl">Fecha ${f.fecha}</div>`;
    f.partidos.forEach(p=>{
      const draw=p.gl===p.gv,lW=p.gl>p.gv,vW=p.gv>p.gl;
      html+=`<div class="match-card">
        <div class="match-team ${lW?'winner':''}" style="opacity:${vW?.6:1}">${p.local}</div>
        <div class="match-score ${draw?'draw':''}">${p.gl} — ${p.gv}</div>
        <div class="match-team right ${vW?'winner':''}" style="opacity:${lW?.6:1}">${p.visitante}</div>
      </div>`;
    });
    html+='</div>';
  });
  return html;
}

function buildSancionados(divNombre){
  const lista=DATA.sancionados[divNombre]||[];
  if(!lista.length) return`<div class="empty">Sin sancionados registrados</div>`;
  let rows='';
  lista.forEach(s=>{
    const numD=isNaN(s.numero)?`<span style="font-size:9px;font-weight:700;color:var(--text3)">${s.numero}</span>`:`<span class="s-num">${s.numero}</span>`;
    rows+=`<tr onmouseover="this.querySelectorAll('td').forEach(t=>t.style.background='var(--surf2)')" onmouseout="this.querySelectorAll('td').forEach(t=>t.style.background='')">
      <td style="text-align:center;padding:10px 12px;border-bottom:1px solid var(--border)">${numD}</td>
      <td style="padding:10px 12px;border-bottom:1px solid var(--border)"><span class="sanc-name">${s.nombre}</span></td>
      <td style="padding:10px 12px;border-bottom:1px solid var(--border);color:var(--text2);font-weight:600">${s.club}</td>
      <td style="padding:10px 12px;border-bottom:1px solid var(--border)"><span class="s-tag ${s.tipo}">${s.sancion}</span></td>
      <td style="text-align:center;padding:10px 12px;border-bottom:1px solid var(--border)">${s.fase2?`<span class="s-fase2">${s.fase2}</span>`:'-'}</td>
    </tr>`;
  });
  return`<div style="overflow-x:auto;border-radius:10px;border:1px solid var(--border)">
    <table class="sanc-table"><thead><tr>
      <th style="text-align:center;width:44px">#</th><th>Jugador / Dirigente</th><th>Club</th>
      <th>Sanción</th><th style="text-align:center">Habilitación</th>
    </tr></thead><tbody>${rows}</tbody></table></div>
  <div style="display:flex;gap:14px;flex-wrap:wrap;padding:10px 4px;font-size:10px;font-weight:600;color:var(--text3)">
    <span><span class="s-tag grave">Grave</span> Indefinido/Años/Meses</span>
    <span><span class="s-tag media">Media</span> 2-3 Fechas</span>
    <span><span class="s-tag leve">Leve</span> 1 Fecha</span>
  </div>`;
}

function buildDivPanel(divNombre){
  const divData=DATA.divisiones[divNombre];
  const color=DC[divNombre]||'#c8102e';
  const meta=divData.meta;
  const totalG=Object.values(divData.teamsData).reduce((a,t)=>a+t.total_goles,0);
  const totalA=Object.values(divData.teamsData).reduce((a,t)=>a+t.total_amarillas,0);
  const totalR=Object.values(divData.teamsData).reduce((a,t)=>a+t.total_rojas+t.total_doble,0);
  const ne=Object.keys(divData.teamsData).length;
  const p=document.getElementById('view-'+divNombre);

  p.innerHTML=`
  <div class="div-hero" style="--accent:${color}">
    <div>
      <h2 class="div-hero-title" style="color:${color}">${divNombre}</h2>
      <div class="div-hero-meta">Fecha ${meta.ultima_fecha} · ${ne} equipos</div>
    </div>
    <div class="div-kpis">
      <div style="text-align:center"><div class="div-kpi-val" style="color:${color}">${totalG}</div><div class="div-kpi-lbl">Goles</div></div>
      <div style="text-align:center"><div class="div-kpi-val">${totalA}</div><div class="div-kpi-lbl">Amarillas</div></div>
      <div style="text-align:center"><div class="div-kpi-val">${totalR}</div><div class="div-kpi-lbl">Rojas/Dbl</div></div>
    </div>
  </div>

  <div class="sub-nav">
    <div class="sub-nav-inner">
      <button class="subn-btn active" data-sub="pos" onclick="selectSub('${divNombre}','pos')"><span class="material-symbols-outlined">format_list_numbered</span>Posiciones</button>
      <button class="subn-btn" data-sub="res" onclick="selectSub('${divNombre}','res')"><span class="material-symbols-outlined">calendar_today</span>Resultados</button>
      <button class="subn-btn" data-sub="stats" onclick="selectSub('${divNombre}','stats')"><span class="material-symbols-outlined">analytics</span>Estadísticas</button>
      <button class="subn-btn" data-sub="sanc" onclick="selectSub('${divNombre}','sanc')"><span class="material-symbols-outlined">gavel</span>Sancionados</button>
    </div>
  </div>

  <div id="sub-${divNombre}-pos" class="sub-panel active">
    <div class="panel">
      <div class="panel-head"><span class="material-symbols-outlined">format_list_numbered</span><span class="panel-head-title">Tabla de Posiciones · ${divNombre}</span></div>
      <div class="panel-body" style="padding:0">${buildTablaPos(divNombre)}</div>
    </div>
  </div>

  <div id="sub-${divNombre}-res" class="sub-panel">
    <div class="panel">
      <div class="panel-head"><span class="material-symbols-outlined">calendar_today</span><span class="panel-head-title">Resultados · ${divNombre}</span></div>
      <div class="panel-body">${buildResultados(divNombre)}</div>
    </div>
  </div>

  <div id="sub-${divNombre}-stats" class="sub-panel">
    <div style="margin-bottom:18px">
      <div class="s-eyebrow" style="color:${color}">Estadísticas</div>
      <div style="font-family:'Barlow Condensed',sans-serif;font-size:22px;font-weight:900;text-transform:uppercase;color:var(--gray)">Selecciona tu equipo para ver sus datos</div>
    </div>
    <div class="team-sel-wrap">
      <span class="team-sel-lbl">Equipo</span>
      <select class="team-sel" id="sel-${divNombre}" onchange="renderTeam('${divNombre}',this.value)" style="border-color:${color}">
        <option value="__global__">📊 Estadística General de la División</option>
      </select>
    </div>
    <div id="dv-global-${divNombre}">
      <div class="grid-3">
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">leaderboard</span><span class="panel-head-title">Top Goleadores</span></div>
          <div class="panel-body"><div id="dg-${divNombre}"></div></div>
        </div>
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">warning</span><span class="panel-head-title">Ranking Amonestaciones</span></div>
          <div class="panel-body"><div id="dt-${divNombre}"></div></div>
        </div>
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined" style="color:var(--green)">volunteer_activism</span><span class="panel-head-title">Fair Play</span></div>
          <div class="panel-body" id="df-${divNombre}"></div>
        </div>
      </div>
    </div>
    <div id="dv-team-${divNombre}" style="display:none"></div>
  </div>

  <div id="sub-${divNombre}-sanc" class="sub-panel">
    <div class="panel">
      <div class="panel-head"><span class="material-symbols-outlined">gavel</span><span class="panel-head-title">Sancionados · ${divNombre}</span></div>
      <div class="panel-body">
        <input type="text" class="search-input" placeholder="Buscar jugador, club o sanción..." oninput="filtrarSanc(this,'sanc-tbl-${divNombre}')">
        ${buildSancionados(divNombre)}
      </div>
    </div>
  </div>`;

  const sel=document.getElementById('sel-'+divNombre);
  Object.keys(divData.teamsData).sort().forEach(t=>{
    const o=document.createElement('option');o.value=t;o.textContent='⚽ '+t;sel.appendChild(o);
  });

  const gd=divData.globalData;
  const maxG2=gd.top_10_goleadores[0]?.goles||1;
  const dgEl=document.getElementById('dg-'+divNombre);
  gd.top_10_goleadores.forEach((pl,i)=>{
    const pct=Math.round(pl.goles/maxG2*100);
    dgEl.innerHTML+=`<div class="scorer-row ${i===0?'top1':i<3?'top3':''}">${rn(i)}<div class="scorer-info">
      <div class="scorer-name">${pl.nombre}</div><div class="scorer-meta" style="color:${color}">${pl.equipo}</div>
      <div class="scorer-bar-wrap"><div class="scorer-bar" style="width:${pct}%;background:${color}"></div></div>
    </div><div class="scorer-val">${pl.goles}</div></div>`;
  });
  const dtEl=document.getElementById('dt-'+divNombre);
  gd.top_10_amonestados.forEach((c,i)=>{
    dtEl.innerHTML+=`<div class="scorer-row">${rn(i)}<div class="scorer-info">
      <div class="scorer-name">${c.nombre}</div><div class="scorer-meta" style="color:${color}">${c.equipo}</div>
    </div><div style="display:flex;gap:2px;flex-wrap:wrap">${cardB(c)}</div></div>`;
  });
  const maxFP2=Math.max(...gd.fair_play.map(f=>f.total_tarjetas))||1;
  const dfEl=document.getElementById('df-'+divNombre);
  gd.fair_play.forEach((eq,i)=>{
    const pct=Math.round(eq.total_tarjetas/maxFP2*100);
    const bc=i<3?'#1a7a3c':i<7?'#b8860b':'#c8102e';
    dfEl.innerHTML+=`<div class="fp-row">${rn(i)}<div class="scorer-info">
      <div class="scorer-name" style="font-size:11px">${eq.equipo}</div>
      <div class="fp-bar-wrap"><div class="fp-bar" style="width:${pct}%;background:${bc}"></div></div>
    </div><div class="fp-val">${eq.total_tarjetas}</div></div>`;
  });
}

function renderTeam(divNombre,equipo){
  const gl=document.getElementById('dv-global-'+divNombre);
  const tm=document.getElementById('dv-team-'+divNombre);
  if(equipo==='__global__'){tm.style.display='none';gl.style.display='block';return;}
  gl.style.display='none';tm.style.display='block';
  const d=DATA.divisiones[divNombre].teamsData[equipo];
  const color=DC[divNombre]||'#c8102e';

  const ultimos=(DATA.ultimos_por_equipo||{})[equipo]||[];
  let formaHTML='<div class="empty">Sin partidos registrados</div>';
  if(ultimos.length){
    formaHTML=`<div class="forma-wrap">`;
    ultimos.forEach(u=>{ formaHTML+=`<div class="forma-match"><span class="forma-pill ${u.resultado}">${u.resultado}</span><span class="forma-score">${u.gf}-${u.gc}</span></div>`; });
    formaHTML+='</div><div style="margin-top:10px">';
    ultimos.forEach(u=>{
      formaHTML+=`<div style="display:flex;align-items:center;gap:10px;padding:6px 10px;border-radius:7px;background:var(--surf2);margin-bottom:3px;font-size:11px">
        <span class="result-pill ${u.resultado}">${u.resultado}</span>
        <span style="font-weight:600;color:var(--text);flex:1">${u.local?'🏠 ':''}vs ${u.rival}</span>
        <span style="font-family:'Barlow Condensed',sans-serif;font-size:15px;font-weight:800;color:var(--gray)">${u.gf}–${u.gc}</span>
        <span style="font-size:9px;color:var(--text4)">F${u.fecha}</span>
      </div>`;
    });
    formaHTML+='</div>';
  }

  // Partidos programados — muestra resultado si ya pasó la fecha/hora, sino el fixture
  const horEq=(DATA.horarios||[]).filter(p=>p.local===equipo||p.visitante===equipo);
  let pxHTML='<div class="empty">Sin partidos programados</div>';
  if(horEq.length){
    pxHTML='';
    const nowEq=new Date();
    horEq.forEach(p=>{
      const horaFmtEq=p.hora==='Sin definir'?'00:00':p.hora.replace('h',':');
      const matchDtEq=new Date(p.fecha_iso+'T'+horaFmtEq+':00');
      const isPastEq=nowEq>matchDtEq;
      const hasResultEq=p.gl!==undefined&&p.gl!==null;
      const lTeam=p.local===equipo;
      if(isPastEq&&hasResultEq){
        const gl=p.gl,gv=p.gv;
        const lw=gl>gv,vw=gv>gl;
        pxHTML+=`<div class="fixture-card resultado-card" style="--accent:${color}">
          <div class="fixture-time"><div style="font-family:'Barlow Condensed',sans-serif;font-size:10px;font-weight:800;letter-spacing:1px;color:var(--green);background:var(--green-light);padding:2px 6px;border-radius:4px">FIN</div><div class="fixture-dia">${p.dia}</div></div>
          <div class="fixture-teams">
            <div class="fixture-team" style="${p.local===equipo?'font-weight:900;color:'+color:lw?'font-weight:700':''}">${p.local}</div>
            <div class="fixture-vs" style="font-family:'Barlow Condensed',sans-serif;font-size:19px;font-weight:900;color:var(--gray);background:linear-gradient(135deg,var(--surf2),var(--surf3));border:1px solid var(--border2);padding:4px 14px;border-radius:8px;min-width:58px;text-align:center;letter-spacing:2px">${gl} – ${gv}</div>
            <div class="fixture-team" style="${p.visitante===equipo?'font-weight:900;color:'+color:vw?'font-weight:700':''}">${p.visitante}</div>
          </div>
          <div class="fixture-cancha"><span class="material-symbols-outlined" style="font-size:13px">location_on</span>${p.cancha}</div>
          ${p.veedor?`<div class="fixture-veedor"><span class="material-symbols-outlined" style="font-size:13px">badge</span>${p.veedor}</div>`:''}
        </div>`;
      } else {
        pxHTML+=`<div class="fixture-card" style="--accent:${color}">
          <div class="fixture-time"><div class="fixture-hora${p.hora==='Sin definir'?' tbd':''}">${p.hora}</div><div class="fixture-dia">${p.dia}</div></div>
          <div class="fixture-teams">
            <div class="fixture-team" style="${p.local===equipo?'font-weight:900;color:'+color:''}">${p.local}</div>
            <div class="fixture-vs">VS</div>
            <div class="fixture-team" style="${p.visitante===equipo?'font-weight:900;color:'+color:''}">${p.visitante}</div>
          </div>
          <div class="fixture-cancha"><span class="material-symbols-outlined" style="font-size:13px">location_on</span>${p.cancha}</div>
          ${p.veedor?`<div class="fixture-veedor"><span class="material-symbols-outlined" style="font-size:13px">badge</span>${p.veedor}</div>`:''}
        </div>`;
      }
    });
  }

  // Próximos rivales — SOLO del Excel de Rivales (solo nombre del rival, sin horario/cancha)
  const rivProx=(DATA.rivalesProximos[divNombre]||{})[equipo]||[];
  let rivalesHTML='<div class="empty">Sin más rivales programados</div>';
  if(rivProx.length){
    rivalesHTML='';
    rivProx.forEach(r=>{
      rivalesHTML+=`<div style="display:flex;align-items:center;gap:10px;padding:10px 12px;border-radius:8px;background:var(--surf2);margin-bottom:5px">
        <span style="font-family:'Barlow Condensed',sans-serif;font-size:11px;font-weight:800;color:var(--text3);background:var(--surf);border:1px solid var(--border);padding:3px 8px;border-radius:4px;min-width:30px;text-align:center">${r.fecha_label}</span>
        <span style="font-family:'Barlow Condensed',sans-serif;font-size:14px;font-weight:800;color:var(--gray);flex:1">vs ${r.rival}</span>
      </div>`;
    });
  }

  const maxG3=d.top_scorers[0]?.goles||1;
  let golesH=d.top_scorers.length?'':'<div class="empty">Sin goles</div>';
  d.top_scorers.forEach((g,i)=>{
    const pct=Math.round(g.goles/maxG3*100);
    golesH+=`<div class="scorer-row ${i===0?'top1':''}">${rn(i)}<div class="scorer-info">
      <div class="scorer-name">${g.nombre}</div>
      <div class="scorer-bar-wrap"><div class="scorer-bar" style="width:${pct}%;background:${color}"></div></div>
    </div><div class="scorer-val">${g.goles}</div></div>`;
  });

  let tarjH=d.top_cards.length?'':'';
  d.top_cards.forEach((c,i)=>{
    tarjH+=`<div class="scorer-row">${rn(i)}<div class="scorer-info"><div class="scorer-name">${c.nombre}</div></div>
      <div style="display:flex;gap:2px;flex-wrap:wrap">${cardB(c)}</div></div>`;
  });
  const disciplinaPanel = d.top_cards.length
    ? `<div class="panel">
         <div class="panel-head"><span class="material-symbols-outlined">warning</span><span class="panel-head-title">${equipo} · Sancionados / Disciplina</span></div>
         <div class="panel-body">${tarjH}</div>
       </div>`
    : `<div class="panel">
         <div class="panel-head"><span class="material-symbols-outlined" style="color:var(--green)">check_circle</span><span class="panel-head-title">${equipo} · Disciplina</span></div>
         <div class="panel-body"><div class="empty">Sin tarjetas registradas — plantel limpio</div></div>
       </div>`;

  document.getElementById('tkpi-'+divNombre)?.remove();
  document.getElementById('tgrid-'+divNombre)?.remove();
  document.getElementById('dv-team-'+divNombre).innerHTML=`
    <div class="div-hero" style="--accent:${color};margin-bottom:16px">
      <div><div class="div-hero-title" style="color:${color};font-size:24px">${equipo}</div><div class="div-hero-meta">${divNombre}</div></div>
    </div>
    <div class="team-kpi-row">
      <div class="team-kpi"><div class="team-kpi-val" style="color:${color}">${d.total_goles}</div><div class="team-kpi-lbl">Goles</div></div>
      <div class="team-kpi"><div class="team-kpi-val">${d.promedio_goles}</div><div class="team-kpi-lbl">Prom. Goles</div></div>
      <div class="team-kpi"><div class="team-kpi-val">${d.total_amarillas}</div><div class="team-kpi-lbl">Amarillas</div></div>
      <div class="team-kpi"><div class="team-kpi-val">${d.total_rojas+d.total_doble}</div><div class="team-kpi-lbl">Rojas/Dbl</div></div>
      <div class="team-kpi"><div class="team-kpi-val">${d.promedio_tarjetas}</div><div class="team-kpi-lbl">Prom. Tarj.</div></div>
    </div>
    <div class="grid-2">
      <div>
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">leaderboard</span><span class="panel-head-title">${equipo} · Goleadores</span></div>
          <div class="panel-body">${golesH}</div>
        </div>
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">sports_score</span><span class="panel-head-title">Últimos 5 Partidos</span></div>
          <div class="panel-body">${formaHTML}</div>
        </div>
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">swap_horiz</span><span class="panel-head-title">Próximos Rivales</span></div>
          <div class="panel-body">${rivalesHTML}</div>
        </div>
      </div>
      <div>
        ${disciplinaPanel}
        <div class="panel">
          <div class="panel-head"><span class="material-symbols-outlined">event</span><span class="panel-head-title">Próximo Partido Programado</span></div>
          <div class="panel-body">${pxHTML}</div>
        </div>
      </div>
    </div>`;
}

buildGlobal();
['1era División','2da División','3era División C1','3era División C2'].forEach(div=>buildDivPanel(div));
</script>
</body>
</html>"""

if __name__=='__main__':
    carpeta=os.path.dirname(os.path.abspath(__file__))
    datos=procesar_todo(carpeta)
    html=generar_html(datos)
    salida=os.path.join(carpeta,'index.html')
    with open(salida,'w',encoding='utf-8') as f: f.write(html)
    safe_print(f"\n[OK] Listo! -> {salida}")
    safe_print(f"   Logo: {'SI' if datos['logo'] else 'NO'}")
    safe_print(f"   Sancionados: {sum(len(v) for v in datos['sancionados'].values())}")
    safe_print(f"   Horarios: {len(datos['horarios'])}")
